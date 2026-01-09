"""Unit tests for Excel document loader (.xlsx, .xls)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch, PropertyMock

import pytest

from ingestion.loader import (
    DocumentParseError,
    load_excel,
    DocumentLoader,
    HANDLERS,
)


class TestExcelBasicLoading:
    """Tests for basic Excel file loading functionality."""

    def test_load_xlsx_single_sheet(self, tmp_path: Path):
        """Test loading simple single-sheet xlsx file."""
        with patch("openpyxl.load_workbook") as mock_load:
            # Create mock workbook with single sheet
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            # Create mock sheet with rows
            mock_sheet = MagicMock()
            mock_cell1 = MagicMock()
            mock_cell1.value = "Name"
            mock_cell2 = MagicMock()
            mock_cell2.value = "Age"

            mock_cell3 = MagicMock()
            mock_cell3.value = "Alice"
            mock_cell4 = MagicMock()
            mock_cell4.value = 30

            mock_sheet.iter_rows.return_value = [
                [mock_cell1, mock_cell2],
                [mock_cell3, mock_cell4],
            ]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "simple.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "[SHEET:Sheet1]" in text
            assert "Name | Age" in text
            assert "Alice | 30" in text
            assert metadata["sheet_count"] == 1
            assert metadata["sheet_names"] == ["Sheet1"]
            assert metadata["total_rows"] == 2

    def test_load_xlsx_multiple_sheets(self, tmp_path: Path):
        """Test loading multi-sheet workbook."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Data", "Summary", "Notes"]
            mock_workbook.properties = None

            # Create different sheets using side_effect for __getitem__
            def get_sheet(name):
                mock_sheet = MagicMock()
                mock_cell = MagicMock()
                mock_cell.value = f"Content from {name}"
                mock_sheet.iter_rows.return_value = [[mock_cell]]
                return mock_sheet

            mock_workbook.__getitem__.side_effect = get_sheet
            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "multi_sheet.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "[SHEET:Data]" in text
            assert "[SHEET:Summary]" in text
            assert "[SHEET:Notes]" in text
            assert "Content from Data" in text
            assert "Content from Summary" in text
            assert "Content from Notes" in text
            assert metadata["sheet_count"] == 3
            assert metadata["sheet_names"] == ["Data", "Summary", "Notes"]

    def test_load_xls_legacy(self, tmp_path: Path):
        """Test loading legacy .xls format (uses same handler)."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["LegacySheet"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cell = MagicMock()
            mock_cell.value = "Legacy data"
            mock_sheet.iter_rows.return_value = [[mock_cell]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xls_path = tmp_path / "legacy.xls"
            xls_path.touch()

            text, metadata = load_excel(xls_path)

            assert "[SHEET:LegacySheet]" in text
            assert "Legacy data" in text
            # Verify .xls is registered in HANDLERS
            assert ".xls" in HANDLERS
            assert HANDLERS[".xls"] == load_excel

    def test_load_excel_empty_workbook(self, tmp_path: Path):
        """Test handling of empty workbook with no data."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["EmptySheet"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_sheet.iter_rows.return_value = []
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "empty.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            # Empty sheet still gets marker
            assert "[SHEET:EmptySheet]" in text
            assert metadata["total_rows"] == 0
            assert metadata["sheet_count"] == 1


class TestExcelContentExtraction:
    """Tests for Excel content extraction behavior."""

    def test_excel_sheet_markers(self, tmp_path: Path):
        """Test that [SHEET:name] markers are present for each sheet."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Finances", "Employees"]
            mock_workbook.properties = None

            def get_sheet(name):
                mock_sheet = MagicMock()
                mock_cell = MagicMock()
                mock_cell.value = f"Data"
                mock_sheet.iter_rows.return_value = [[mock_cell]]
                return mock_sheet

            mock_workbook.__getitem__.side_effect = get_sheet
            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "sheets.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "[SHEET:Finances]" in text
            assert "[SHEET:Employees]" in text
            # Markers should be at beginning of each sheet section
            assert text.index("[SHEET:Finances]") < text.index("[SHEET:Employees]")

    def test_excel_row_formatting(self, tmp_path: Path):
        """Test that rows use pipe-separated format."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cells = [MagicMock() for _ in range(4)]
            mock_cells[0].value = "Col1"
            mock_cells[1].value = "Col2"
            mock_cells[2].value = "Col3"
            mock_cells[3].value = "Col4"

            mock_sheet.iter_rows.return_value = [mock_cells]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "columns.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "Col1 | Col2 | Col3 | Col4" in text

    def test_excel_merged_cells(self, tmp_path: Path):
        """Test handling of merged cells (values appear in first cell only)."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            # Merged cells: first cell has value, others are None
            mock_cell1 = MagicMock()
            mock_cell1.value = "Merged Header"
            mock_cell2 = MagicMock()
            mock_cell2.value = None  # Merged with cell1
            mock_cell3 = MagicMock()
            mock_cell3.value = None  # Merged with cell1

            mock_sheet.iter_rows.return_value = [[mock_cell1, mock_cell2, mock_cell3]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "merged.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            # Should handle None values gracefully
            assert "Merged Header" in text
            assert metadata["total_rows"] == 1

    def test_excel_formulas(self, tmp_path: Path):
        """Test that formula results (not formulas) are extracted."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            # With data_only=True, formulas return their computed values
            mock_cell1 = MagicMock()
            mock_cell1.value = 100  # Result of =SUM(A1:A10)
            mock_cell2 = MagicMock()
            mock_cell2.value = 42  # Result of =A1*B1

            mock_sheet.iter_rows.return_value = [[mock_cell1, mock_cell2]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "formulas.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "100 | 42" in text
            # Verify data_only=True was passed to load_workbook
            mock_load.assert_called_once()
            call_kwargs = mock_load.call_args[1]
            assert call_kwargs.get("data_only") is True

    def test_excel_empty_rows(self, tmp_path: Path):
        """Test handling of sparse data with empty rows."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()

            # Row 1: has data
            mock_cell1 = MagicMock()
            mock_cell1.value = "Header"

            # Row 2: empty (all None)
            mock_cell2 = MagicMock()
            mock_cell2.value = None
            mock_cell3 = MagicMock()
            mock_cell3.value = None

            # Row 3: has data
            mock_cell4 = MagicMock()
            mock_cell4.value = "Data"

            mock_sheet.iter_rows.return_value = [
                [mock_cell1],
                [mock_cell2, mock_cell3],  # Empty row
                [mock_cell4],
            ]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "sparse.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "Header" in text
            assert "Data" in text
            # Empty rows should be skipped
            assert metadata["total_rows"] == 2


class TestExcelMetadataExtraction:
    """Tests for Excel metadata extraction."""

    def test_excel_metadata_extraction(self, tmp_path: Path):
        """Test extraction of all metadata fields from document properties."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]

            mock_props = MagicMock()
            mock_props.title = "Financial Report 2024"
            mock_props.creator = "John Smith"
            mock_props.created = datetime(2024, 1, 15, 10, 30, 0)
            mock_props.modified = datetime(2024, 3, 20, 14, 45, 0)
            mock_props.subject = "Q4 Analysis"
            mock_props.keywords = "finance, quarterly, report"

            mock_workbook.properties = mock_props

            mock_sheet = MagicMock()
            mock_cell = MagicMock()
            mock_cell.value = "Data"
            mock_sheet.iter_rows.return_value = [[mock_cell]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "report.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert metadata["title"] == "Financial Report 2024"
            assert metadata["author"] == "John Smith"
            assert "2024-01-15" in metadata["creation_date"]
            assert "2024-03-20" in metadata["modification_date"]
            assert metadata["subject"] == "Q4 Analysis"
            assert metadata["keywords"] == "finance, quarterly, report"

    def test_excel_sheet_names_metadata(self, tmp_path: Path):
        """Test that sheet_names list is correctly populated in metadata."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sales", "Expenses", "Profit", "Forecast"]
            mock_workbook.properties = None

            def get_sheet(name):
                mock_sheet = MagicMock()
                mock_sheet.iter_rows.return_value = []
                return mock_sheet

            mock_workbook.__getitem__.side_effect = get_sheet
            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "multi.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert metadata["sheet_names"] == ["Sales", "Expenses", "Profit", "Forecast"]
            assert metadata["sheet_count"] == 4

    def test_excel_no_metadata(self, tmp_path: Path):
        """Test handling of files without document properties."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None  # No properties

            mock_sheet = MagicMock()
            mock_cell = MagicMock()
            mock_cell.value = "Data"
            mock_sheet.iter_rows.return_value = [[mock_cell]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "no_metadata.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            # Should use filename as title fallback
            assert metadata["title"] == "no_metadata"
            # Should not have author, creation_date, etc.
            assert "author" not in metadata
            assert "creation_date" not in metadata

    def test_excel_title_fallback_to_filename(self, tmp_path: Path):
        """Test that filename is used as title when not in properties."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]

            mock_props = MagicMock()
            mock_props.title = None  # No title
            mock_props.creator = None
            mock_props.created = None
            mock_props.modified = None
            mock_props.subject = None
            mock_props.keywords = None

            mock_workbook.properties = mock_props

            mock_sheet = MagicMock()
            mock_sheet.iter_rows.return_value = []
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "my_spreadsheet.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert metadata["title"] == "my_spreadsheet"


class TestExcelErrorHandling:
    """Tests for Excel error handling scenarios."""

    def test_excel_corrupted_file(self, tmp_path: Path):
        """Test that corrupted Excel files raise DocumentParseError."""
        from openpyxl.utils.exceptions import InvalidFileException

        with patch("openpyxl.load_workbook") as mock_load:
            mock_load.side_effect = InvalidFileException("File is not a valid Excel file")

            xlsx_path = tmp_path / "corrupted.xlsx"
            xlsx_path.touch()

            with pytest.raises(DocumentParseError) as exc_info:
                load_excel(xlsx_path)

            assert "Corrupted Excel file" in str(exc_info.value)
            assert exc_info.value.path == xlsx_path

    def test_excel_password_protected(self, tmp_path: Path):
        """Test handling of password-protected/encrypted workbooks."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_load.side_effect = Exception("Workbook is encrypted")

            xlsx_path = tmp_path / "protected.xlsx"
            xlsx_path.touch()

            with pytest.raises(DocumentParseError) as exc_info:
                load_excel(xlsx_path)

            assert "Password-protected" in str(exc_info.value)

    def test_excel_password_protected_variant(self, tmp_path: Path):
        """Test handling of password error message variant."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_load.side_effect = Exception("Cannot open password-protected file")

            xlsx_path = tmp_path / "locked.xlsx"
            xlsx_path.touch()

            with pytest.raises(DocumentParseError) as exc_info:
                load_excel(xlsx_path)

            assert "Password-protected" in str(exc_info.value)

    def test_excel_general_exception(self, tmp_path: Path):
        """Test handling of general exceptions during Excel loading."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_load.side_effect = RuntimeError("Unexpected error")

            xlsx_path = tmp_path / "error.xlsx"
            xlsx_path.touch()

            with pytest.raises(DocumentParseError) as exc_info:
                load_excel(xlsx_path)

            assert "Excel read error" in str(exc_info.value)

    def test_excel_unsupported_extension(self, tmp_path: Path):
        """Test that unsupported extensions are not handled."""
        # .xlsm (macro-enabled) is not registered
        assert ".xlsm" not in HANDLERS

        # .xlsx and .xls should be registered
        assert ".xlsx" in HANDLERS
        assert ".xls" in HANDLERS
        assert HANDLERS[".xlsx"] == load_excel
        assert HANDLERS[".xls"] == load_excel


class TestExcelIntegration:
    """Integration tests for Excel loading through the pipeline."""

    def test_excel_through_pipeline(self, tmp_path: Path):
        """Test full ingestion of Excel file through DocumentLoader."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Data"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cell1 = MagicMock()
            mock_cell1.value = "Product"
            mock_cell2 = MagicMock()
            mock_cell2.value = "Price"

            mock_cell3 = MagicMock()
            mock_cell3.value = "Widget"
            mock_cell4 = MagicMock()
            mock_cell4.value = 19.99

            mock_sheet.iter_rows.return_value = [
                [mock_cell1, mock_cell2],
                [mock_cell3, mock_cell4],
            ]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "products.xlsx"
            xlsx_path.write_bytes(b"fake excel content")

            loader = DocumentLoader(input_root=tmp_path)
            document, content_hash = loader.load(xlsx_path)

            assert document.source_type == "xlsx"
            assert "[SHEET:Data]" in document.text
            assert "Product | Price" in document.text
            assert "Widget | 19.99" in document.text
            assert document.metadata["sheet_count"] == 1
            assert content_hash is not None

    def test_excel_through_pipeline_xls(self, tmp_path: Path):
        """Test ingestion of .xls file through DocumentLoader."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Legacy"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cell = MagicMock()
            mock_cell.value = "Old data"
            mock_sheet.iter_rows.return_value = [[mock_cell]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xls_path = tmp_path / "legacy.xls"
            xls_path.write_bytes(b"fake xls content")

            loader = DocumentLoader(input_root=tmp_path)
            document, content_hash = loader.load(xls_path)

            assert document.source_type == "xls"
            assert "[SHEET:Legacy]" in document.text

    def test_excel_chunking(self, tmp_path: Path):
        """Test that Excel content can be chunked correctly."""
        from ingestion.chunker import chunk_document
        from ingestion.models import Document

        # Create a document with substantial content
        excel_content = "[SHEET:Sheet1]\n"
        excel_content += "Header1 | Header2 | Header3\n"
        for i in range(100):
            excel_content += f"Row{i} | Value{i} | Description of item {i} with more text to make it longer\n"

        mock_document = Document(
            doc_id="test-excel",
            path=tmp_path / "test.xlsx",
            source_type="xlsx",
            text=excel_content,
            metadata={
                "sheet_count": 1,
                "sheet_names": ["Sheet1"],
                "total_rows": 101,
            },
        )

        chunks = chunk_document(mock_document)

        assert len(chunks) > 0
        # First chunk should contain sheet marker
        assert "[SHEET:Sheet1]" in chunks[0].text
        # Chunks should have proper IDs
        assert all(c.chunk_id.startswith("test-excel::chunk-") for c in chunks)

    def test_excel_metadata_preserved_through_pipeline(self, tmp_path: Path):
        """Test that Excel metadata is preserved through DocumentLoader."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Q1", "Q2"]

            mock_props = MagicMock()
            mock_props.title = "Quarterly Report"
            mock_props.creator = "Finance Team"
            mock_props.created = datetime(2024, 4, 1, 9, 0, 0)
            mock_props.modified = None
            mock_props.subject = None
            mock_props.keywords = None

            mock_workbook.properties = mock_props

            def get_sheet(name):
                mock_sheet = MagicMock()
                mock_cell = MagicMock()
                mock_cell.value = f"{name} data"
                mock_sheet.iter_rows.return_value = [[mock_cell]]
                return mock_sheet

            mock_workbook.__getitem__.side_effect = get_sheet
            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "quarterly.xlsx"
            xlsx_path.write_bytes(b"fake excel")

            loader = DocumentLoader(input_root=tmp_path)
            document, _ = loader.load(xlsx_path)

            assert document.metadata["title"] == "Quarterly Report"
            assert document.metadata["author"] == "Finance Team"
            assert "2024-04-01" in document.metadata["creation_date"]
            assert document.metadata["sheet_names"] == ["Q1", "Q2"]
            assert document.metadata["sheet_count"] == 2


class TestExcelEdgeCases:
    """Tests for Excel edge cases and boundary conditions."""

    def test_excel_special_characters_in_sheet_name(self, tmp_path: Path):
        """Test handling of special characters in sheet names."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Q1 - 2024", "Data (Raw)", "Sheet #1"]
            mock_workbook.properties = None

            def get_sheet(name):
                mock_sheet = MagicMock()
                mock_cell = MagicMock()
                mock_cell.value = "Data"
                mock_sheet.iter_rows.return_value = [[mock_cell]]
                return mock_sheet

            mock_workbook.__getitem__.side_effect = get_sheet
            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "special_names.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "[SHEET:Q1 - 2024]" in text
            assert "[SHEET:Data (Raw)]" in text
            assert "[SHEET:Sheet #1]" in text

    def test_excel_very_wide_row(self, tmp_path: Path):
        """Test handling of rows with many columns."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Wide"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            # Create row with 50 columns
            mock_cells = []
            for i in range(50):
                mock_cell = MagicMock()
                mock_cell.value = f"Col{i}"
                mock_cells.append(mock_cell)

            mock_sheet.iter_rows.return_value = [mock_cells]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "wide.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert metadata["total_columns"] == 50
            # All columns should be pipe-separated
            assert text.count("|") == 49  # 50 columns = 49 separators

    def test_excel_numeric_values_converted_to_string(self, tmp_path: Path):
        """Test that numeric values are properly converted to strings."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Numbers"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cell1 = MagicMock()
            mock_cell1.value = 12345
            mock_cell2 = MagicMock()
            mock_cell2.value = 99.99
            mock_cell3 = MagicMock()
            mock_cell3.value = -42

            mock_sheet.iter_rows.return_value = [[mock_cell1, mock_cell2, mock_cell3]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "numbers.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "12345 | 99.99 | -42" in text

    def test_excel_whitespace_trimmed(self, tmp_path: Path):
        """Test that cell values have whitespace trimmed."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_cell1 = MagicMock()
            mock_cell1.value = "  padded  "
            mock_cell2 = MagicMock()
            mock_cell2.value = "\ttabbed\t"

            mock_sheet.iter_rows.return_value = [[mock_cell1, mock_cell2]]
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "whitespace.xlsx"
            xlsx_path.touch()

            text, metadata = load_excel(xlsx_path)

            assert "padded | tabbed" in text

    def test_excel_workbook_closed_after_reading(self, tmp_path: Path):
        """Test that workbook is properly closed after reading."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_sheet.iter_rows.return_value = []
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "test.xlsx"
            xlsx_path.touch()

            load_excel(xlsx_path)

            # Verify workbook.close() was called
            mock_workbook.close.assert_called_once()

    def test_excel_read_only_mode(self, tmp_path: Path):
        """Test that workbook is opened in read-only mode for performance."""
        with patch("openpyxl.load_workbook") as mock_load:
            mock_workbook = MagicMock()
            mock_workbook.sheetnames = ["Sheet1"]
            mock_workbook.properties = None

            mock_sheet = MagicMock()
            mock_sheet.iter_rows.return_value = []
            mock_workbook.__getitem__ = lambda self, name: mock_sheet

            mock_load.return_value = mock_workbook

            xlsx_path = tmp_path / "test.xlsx"
            xlsx_path.touch()

            load_excel(xlsx_path)

            # Verify read_only=True was passed
            call_kwargs = mock_load.call_args[1]
            assert call_kwargs.get("read_only") is True
